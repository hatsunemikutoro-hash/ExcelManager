# =============================================================================
# EXCEL MANAGER — core.py
# Responsabilidade: lógica de negócio pura, sem formatação de UI.
# Todas as funções retornam dados brutos ou levantam exceções tipadas.
# =============================================================================

import csv
import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook


# =============================================================================
# SEÇÃO 1 — HELPERS INTERNOS
# =============================================================================

def _carregar_workbook(arquivo: str, data_only: bool = False, read_only: bool = False):
    """
    Carrega um workbook detectando automaticamente .xlsx e .xls legado.

    Raises:
        FileNotFoundError : Arquivo não encontrado.
        PermissionError   : Arquivo aberto em outro programa.
        ValueError        : Formato de arquivo não suportado.
    """
    ext = Path(arquivo).suffix.lower()

    if ext == ".xls":
        # xlrd só lê .xls — convertemos para um Workbook openpyxl em memória
        try:
            import xlrd
            book = xlrd.open_workbook(arquivo)
            wb = Workbook()
            ws = wb.active
            sheet = book.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    ws.cell(row=row_idx + 1, column=col_idx + 1,
                            value=sheet.cell_value(row_idx, col_idx))
            return wb
        except ImportError:
            raise ImportError("Instale 'xlrd' para suporte a arquivos .xls legados.")
        except Exception as e:
            raise Exception(f"Erro ao ler arquivo .xls: {e}")

    elif ext == ".xlsx":
        try:
            return load_workbook(arquivo, data_only=data_only, read_only=read_only)
        except PermissionError:
            raise PermissionError(
                "O arquivo está sendo usado por outro programa.\n"
                "Feche o Excel (ou outro programa) e tente novamente."
            )
        except FileNotFoundError:
            raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")
        except Exception as e:
            raise Exception(f"Erro ao abrir o arquivo: {e}")
    else:
        raise ValueError(f"Formato '{ext}' não suportado. Use .xlsx ou .xls.")


# =============================================================================
# SEÇÃO 2 — CRIAÇÃO E ESCRITA
# =============================================================================

def criarPlanilha(lista: list, colunas: int, destino: str) -> str:
    """
    Cria um novo arquivo .xlsx a partir de uma lista plana de valores.

    Args:
        lista   : Itens a serem escritos célula a célula, da esquerda para direita.
        colunas : Quantidade de colunas antes de quebrar para a próxima linha.
        destino : Caminho completo onde o arquivo será salvo (incluindo .xlsx).

    Returns:
        Caminho completo do arquivo salvo.
    """
    if colunas <= 0:
        raise ValueError("O número de colunas deve ser maior que zero.")

    wb = Workbook()
    ws = wb.active

    linha, coluna = 1, 1
    for item in lista:
        ws.cell(row=linha, column=coluna, value=item)
        coluna += 1
        if coluna > colunas:
            coluna = 1
            linha += 1

    try:
        wb.save(destino)
    except PermissionError:
        raise PermissionError(
            "Não foi possível salvar. Feche o arquivo se ele já estiver aberto."
        )
    return destino


def preencherPlanilha(path: str, lista: list, colunas: int, linha_inicial: int = 2) -> str:
    """
    Injeta dados em uma planilha existente a partir de uma linha específica.

    Args:
        path          : Caminho do arquivo .xlsx de destino.
        lista         : Dados a serem inseridos.
        colunas       : Largura da grade (colunas por linha).
        linha_inicial : Linha a partir da qual os dados serão escritos.

    Returns:
        Caminho do arquivo salvo.
    """
    if colunas <= 0:
        raise ValueError("O número de colunas deve ser maior que zero.")

    wb = _carregar_workbook(path)
    ws = wb.active

    linha, coluna = linha_inicial, 1
    for item in lista:
        ws.cell(row=linha, column=coluna, value=item)
        coluna += 1
        if coluna > colunas:
            coluna = 1
            linha += 1

    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError(
            "Não foi possível salvar. Feche o arquivo se ele já estiver aberto."
        )
    return path


def mesclarPlanilhas(arquivos: list[str], destino: str) -> tuple[str, int]:
    """
    Mescla múltiplos arquivos .xlsx/.xls em um único arquivo.
    Os cabeçalhos são lidos do primeiro arquivo; os demais apenas appendam dados.

    Args:
        arquivos : Lista de caminhos de arquivos Excel.
        destino  : Caminho do arquivo de saída.

    Returns:
        Tupla (caminho_destino, total_linhas_mescladas).
    """
    if len(arquivos) < 2:
        raise ValueError("Selecione ao menos 2 arquivos para mesclar.")

    wb_saida = Workbook()
    ws_saida = wb_saida.active
    total_linhas = 0
    primeiro = True

    for arquivo in arquivos:
        wb = _carregar_workbook(arquivo, data_only=True)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Pula o cabeçalho nos arquivos subsequentes
            if not primeiro and i == 0:
                continue
            ws_saida.append(list(row))
            total_linhas += 1

        primeiro = False

    try:
        wb_saida.save(destino)
    except PermissionError:
        raise PermissionError("Feche o arquivo de destino antes de salvar.")

    return destino, total_linhas


# =============================================================================
# SEÇÃO 3 — LEITURA E INSPEÇÃO
# =============================================================================

def contarLinhas(arquivo: str) -> int:
    """
    Retorna o índice da última linha que contém algum dado (ignora linhas vazias no fim).
    """
    wb = _carregar_workbook(arquivo)
    ws = wb.active

    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row=row, column=col).value is not None
               for col in range(1, ws.max_column + 1)):
            return row
    return 0


def lerCabecalhos(arquivo: str) -> list[str]:
    """
    Lê e retorna os cabeçalhos (primeira linha) da planilha ativa.

    Returns:
        Lista de strings com os valores da primeira linha (sem nulos).
    """
    wb = _carregar_workbook(arquivo, data_only=True)
    ws = wb.active

    try:
        primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        raise ValueError("A planilha está completamente vazia.")

    cabecalhos = [str(item).strip() for item in primeira_linha if item is not None]
    if not cabecalhos:
        raise ValueError("Nenhum cabeçalho encontrado na primeira linha.")

    return cabecalhos


def buscarRegistro(arquivo: str, filtro: str) -> dict | None:
    """
    Busca a primeira célula cujo valor contém o termo informado (case-insensitive).

    Returns:
        Dicionário com 'valor', 'linha' e 'coluna' se encontrado, ou None.
    """
    wb = _carregar_workbook(arquivo)
    ws = wb.active

    for linha in ws.iter_rows():
        for celula in linha:
            valor = celula.value
            if valor is not None and filtro.lower() in str(valor).lower():
                return {
                    "valor": str(valor),
                    "linha": celula.row,
                    "coluna": celula.column_letter,
                }
    return None


def listarConteudo(arquivo: str) -> list[tuple]:
    """
    Retorna todas as linhas da planilha ativa como lista de tuplas.
    """
    wb = _carregar_workbook(arquivo, data_only=True)
    ws = wb.active
    return [row for row in ws.iter_rows(values_only=True)]


def listarAbas(arquivo: str) -> list[str]:
    """
    Retorna os nomes de todas as abas (sheets) do arquivo.
    """
    wb = _carregar_workbook(arquivo, read_only=True)
    return wb.sheetnames


def previewDados(lista: list, colunas: int) -> list[list]:
    """
    Simula como os dados ficarão organizados na planilha, sem salvar nada.

    Returns:
        Lista de listas representando as linhas da planilha.
    """
    if colunas <= 0:
        raise ValueError("Número de colunas deve ser maior que zero.")

    grade, linha_atual = [], []
    for item in lista:
        linha_atual.append(item)
        if len(linha_atual) == colunas:
            grade.append(linha_atual)
            linha_atual = []
    if linha_atual:
        grade.append(linha_atual)
    return grade


# =============================================================================
# SEÇÃO 4 — EXPORTAÇÃO
# =============================================================================

def exportarCSV(arquivo_xlsx: str, destino_csv: str, separador: str = ",") -> str:
    """
    Exporta a planilha ativa de um .xlsx para um arquivo .csv.

    Args:
        arquivo_xlsx : Caminho do arquivo Excel de origem.
        destino_csv  : Caminho do arquivo CSV de destino.
        separador    : Caractere separador (padrão: vírgula).

    Returns:
        Caminho do CSV gerado.
    """
    linhas = listarConteudo(arquivo_xlsx)

    with open(destino_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=separador)
        for linha in linhas:
            writer.writerow([v if v is not None else "" for v in linha])

    return destino_csv


# =============================================================================
# SEÇÃO 5 — HISTÓRICO DE OPERAÇÕES
# =============================================================================

HISTORICO_PATH = Path.home() / ".excel_manager_historico.json"


def _carregar_historico() -> list[dict]:
    """Carrega o histórico de operações do disco."""
    if not HISTORICO_PATH.exists():
        return []
    try:
        with open(HISTORICO_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def registrarHistorico(operacao: str, detalhe: str):
    """
    Registra uma operação no histórico persistente.

    Args:
        operacao : Nome da operação (ex: "Criar", "Preencher", "Exportar CSV").
        detalhe  : Informação complementar (ex: nome do arquivo, quantidade de linhas).
    """
    historico = _carregar_historico()
    historico.insert(0, {
        "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "operacao": operacao,
        "detalhe": detalhe,
    })
    # Mantém apenas os últimos 100 registros
    historico = historico[:100]

    try:
        with open(HISTORICO_PATH, "w", encoding="utf-8") as f:
            json.dump(historico, f, ensure_ascii=False, indent=2)
    except Exception:
        pass  # Falha silenciosa — histórico não é crítico


def obterHistorico() -> list[dict]:
    """Retorna a lista de operações registradas (mais recente primeiro)."""
    return _carregar_historico()


def limparHistorico():
    """Apaga todo o histórico de operações."""
    try:
        HISTORICO_PATH.unlink(missing_ok=True)
    except Exception:
        pass